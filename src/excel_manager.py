"""
Módulo de Gerenciamento da Base de Dados em Excel com Categorização por Famílias,
Fluxo Enxuto de Desembolso e Resumos com Semanas Resetadas a Cada Mês.
"""

import os
import shutil
import json
import time
import re
from datetime import datetime, date, timedelta
from typing import List, Dict, Any, Optional
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 1. DEFINIÇÃO DAS COLUNAS DA BASE OPERACIONAL (ABA 1)
COLUMNS_DEFINITION = [
    {"key": "numero_pedido", "title": "Nº Pedido", "width": 14, "align": "center", "format": "@"},
    {"key": "data_pedido", "title": "Data do Pedido", "width": 15, "align": "center", "format": "@"},
    {"key": "fornecedor_nome", "title": "Fornecedor", "width": 36, "align": "left", "format": "@"},
    {"key": "fornecedor_cnpj", "title": "CNPJ Fornecedor", "width": 22, "align": "center", "format": "@"},
    {"key": "codigo_insumo", "title": "Cód. Insumo", "width": 14, "align": "center", "format": "@"},
    {"key": "familia_insumo", "title": "Família Sienge", "width": 34, "align": "left", "format": "@"},
    {"key": "descricao_material", "title": "Descrição do Material / Insumo", "width": 48, "align": "left", "format": "@"},
    {"key": "quantidade", "title": "Quantidade", "width": 16, "align": "right", "format": "#,##0.0000"},
    {"key": "unidade", "title": "Unid.", "width": 10, "align": "center", "format": "@"},
    {"key": "preco_unitario", "title": "Pr. Unitário (R$)", "width": 18, "align": "right", "format": "R$ #,##0.0000"},
    {"key": "preco_total_item", "title": "Total Item (R$)", "width": 18, "align": "right", "format": "R$ #,##0.00"},
    {"key": "data_entrega_prevista", "title": "Entrega Prevista", "width": 16, "align": "center", "format": "@"},
    {"key": "condicao_pagamento", "title": "Condição Pagamento", "width": 24, "align": "center", "format": "@"},
    {"key": "situacao_prazo", "title": "Previsão / Prazo", "width": 22, "align": "center", "format": "@"},
    {"key": "observacoes_sienge", "title": "Observações Sienge", "width": 40, "align": "left", "format": "@"},
    {"key": "arquivo_origem", "title": "Arquivo Fonte", "width": 25, "align": "left", "format": "@"},
    {"key": "ultima_atualizacao", "title": "Última Atualização", "width": 20, "align": "center", "format": "@"},
    {"key": "valor_frete", "title": "Frete Pedido (R$)", "width": 18, "align": "right", "format": "R$ #,##0.00"},
    {"key": "valor_total_pedido", "title": "Total Pedido Sienge (R$)", "width": 22, "align": "right", "format": "R$ #,##0.00"}
]


# 2. DEFINIÇÃO DAS COLUNAS ENXUTAS DO FLUXO DE DESEMBOLSO (ABA 2)
COLUMNS_DESEMBOLSO_ENXUTO = [
    {"key": "numero_pedido", "title": "Nº Pedido", "width": 14, "align": "center", "format": "@"},
    {"key": "fornecedor", "title": "Fornecedor", "width": 34, "align": "left", "format": "@"},
    {"key": "cnpj", "title": "CNPJ", "width": 20, "align": "center", "format": "@"},
    {"key": "familia", "title": "Família / Grupo", "width": 30, "align": "left", "format": "@"},
    {"key": "data_entrega", "title": "Data Entrega Combinada", "width": 22, "align": "center", "format": "@"},
    {"key": "condicao_original", "title": "Condição de Pagamento", "width": 24, "align": "center", "format": "@"},
    {"key": "parcela", "title": "Parcela", "width": 12, "align": "center", "format": "@"},
    {"key": "data_vencimento", "title": "Previsão de Pagamento", "width": 22, "align": "center", "format": "@"},
    {"key": "semana_ano", "title": "Semana Prevista", "width": 28, "align": "center", "format": "@"},
    {"key": "mes_ano", "title": "Mês Previsto", "width": 18, "align": "center", "format": "@"},
    {"key": "valor_parcela", "title": "Valor da Parcela (R$)", "width": 22, "align": "right", "format": "R$ #,##0.00"}
]


NOMES_MESES = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]


CANONICAL_FAMILIES = {
    "INSTAL.HIDRAULICA": "INSTALAÇÕES HIDRÁULICAS",
    "INST. HIDRÁULICAS": "INSTALAÇÕES HIDRÁULICAS",
    "INST. HIDRAULICAS": "INSTALAÇÕES HIDRÁULICAS",
    "INST HIDRÁULICAS": "INSTALAÇÕES HIDRÁULICAS",
    "INST HIDRAULICAS": "INSTALAÇÕES HIDRÁULICAS",
    "INSTALAÇÕES HIDRÁULICAS": "INSTALAÇÕES HIDRÁULICAS",
    "INSTALACOES HIDRAULICAS": "INSTALAÇÕES HIDRÁULICAS",
    "INSTALAÇÕES ELÉTRICAS-TELEFÔNICAS": "INSTALAÇÕES ELÉTRICAS",
    "INSTALACOES ELETRICAS-TELEFONICAS": "INSTALAÇÕES ELÉTRICAS",
    "INSTALAÇÕES ELÉTRICAS": "INSTALAÇÕES ELÉTRICAS",
    "INSTALACOES ELETRICAS": "INSTALAÇÕES ELÉTRICAS",
    "INST. ELÉTRICAS": "INSTALAÇÕES ELÉTRICAS",
    "INST. ELETRICAS": "INSTALAÇÕES ELÉTRICAS",
    "INST ELÉTRICAS": "INSTALAÇÕES ELÉTRICAS",
    "INST ELETRICAS": "INSTALAÇÕES ELÉTRICAS",
    "ENERGIA ELÉTRICA - L.T. - S.E.": "INSTALAÇÕES ELÉTRICAS",
    "ENERGIA ELETRICA - L.T. - S.E.": "INSTALAÇÕES ELÉTRICAS",
    "PRODUTOS METALICOS": "PRODUTOS METÁLICOS",
    "PRODUTOS METÁLICOS": "PRODUTOS METÁLICOS",
    "EPI'S E EPC'S": "EPI'S E EPC'S",
    "EPI'S e EPC'S": "EPI'S E EPC'S",
    "EPIS E EPCS": "EPI'S E EPC'S",
    "EQUIPAMENTOS DE PROTECAO": "EPI'S E EPC'S",
    "IMPERMEABILIZANTE-ISOLAMENTO-JUNTAS": "IMPERMEABILIZANTES",
    "IMPERMEABILIZANTES": "IMPERMEABILIZANTES",
    "VEDAÇÃO": "IMPERMEABILIZANTES",
    "VEDACAO": "IMPERMEABILIZANTES",
    "LOUÇAS, METAIS E ACESSORIOS": "LOUÇAS E METAIS",
    "LOUCAS, METAIS E ACESSORIOS": "LOUÇAS E METAIS",
    "LOUÇAS E METAIS": "LOUÇAS E METAIS",
    "LOUCAS E METAIS": "LOUÇAS E METAIS",
    "INST. DE INCENDIO": "INSTALAÇÕES DE INCÊNDIO",
    "INST. DE INCÊNDIO": "INSTALAÇÕES DE INCÊNDIO",
    "INST DE INCENDIO": "INSTALAÇÕES DE INCÊNDIO",
    "INST DE INCÊNDIO": "INSTALAÇÕES DE INCÊNDIO",
    "INSTALAÇÕES DE INCÊNDIO": "INSTALAÇÕES DE INCÊNDIO",
    "INST. DE GÁS": "INSTALAÇÕES DE GÁS",
    "INST. DE GAS": "INSTALAÇÕES DE GÁS",
    "INST DE GÁS": "INSTALAÇÕES DE GÁS",
    "INST DE GAS": "INSTALAÇÕES DE GÁS",
    "INSTALAÇÕES DE GÁS": "INSTALAÇÕES DE GÁS",
    "BLOCOS,TIJOLOS E ELEMENTOS VAZADOS": "BLOCOS E TIJOLOS",
    "BLOCOS E TIJOLOS": "BLOCOS E TIJOLOS",
    "AGREGADOS E AGLOMERANTES": "AGREGADOS E AGLOMERANTES",
    "ARGAMASSAS": "ARGAMASSAS",
    "MADEIRA": "MADEIRA",
    "FERRAMENTAS": "FERRAMENTAS",
    "FERRAGENS": "FERRAGENS",
    "TINTAS": "TINTAS",
    "MATERIAL DE LIMPEZA": "MATERIAL DE LIMPEZA",
    "MATERIAL DE EXPEDIENTE": "MATERIAL DE EXPEDIENTE",
    "PAVIMENTAÇÃO E DRENAGEM": "PAVIMENTAÇÃO E DRENAGEM",
    "PAVIMENTACAO E DRENAGEM": "PAVIMENTAÇÃO E DRENAGEM",
    "SERVIÇOS EMPREITADOS": "SERVIÇOS EMPREITADOS",
    "SERVICOS EMPREITADOS": "SERVIÇOS EMPREITADOS",
    "EQUIPAMENTOS PRÓPRIOS": "EQUIPAMENTOS",
    "EQUIPAMENTOS PROPRIOS": "EQUIPAMENTOS",
    "EQUIPAMENTOS": "EQUIPAMENTOS",
    "PRODUTOS INDUSTRIALIZADOS": "MATERIAIS AUXILIARES",
    "SINALIZAÇÃO": "EPI'S E EPC'S",
    "SINALIZACAO": "EPI'S E EPC'S",
    "REVESTIMENTO": "FERRAMENTAS",
    "DIVERSOS": "MATERIAIS AUXILIARES",
    "MATERIAIS AUXILIARES": "MATERIAIS AUXILIARES"
}


def classify_item(desc: str, cod: str = "", current_fam: str = "") -> str:
    d = (desc or "").upper()
    c = (cod or "").strip()

    # 1. Produtos Metálicos e Serralheria
    if any(k in d for k in ["CHAPA GALVANIZADA", "BARRA CHATA", "BARRA ROSCADA", "ARAME GALVANIZADO", "ARAME RECOZIDO", "VERGALHAO", "VERGALHÃO", "CA-50", "CA-60", "ESTRIBO", "CORDOALHA", "CUNHA", "ANCORAGEM", "PERFIL DE ALUMINIO", "PERFIL DE ALUMÍNIO", "MAO FRANCESA", "MÃO FRANCESA", "JAQUETA E CONE", "CHUMBADOR", "PARAFUSO SEXTAVADO", "PARAFUSO FRANCES", "PORCA SEXTAVADA", "ARRUELA", "ELETRODO", "SOLDA ESTANHO", "PASTA PARA SOLDAR", "PINO DE ACO", "PINO DE AÇO", "FINCAPINO", "CANTONEIRA"]):
        return "PRODUTOS METÁLICOS"

    # 2. Instalações Hidráulicas
    if any(k in d for k in ["TUBO PVC", "TUBO COBRE", "TUBO PPR", "TUBO CPVC", "CONEXAO PVC", "CONEXÃO PVC", "JOELHO PVC", "JOELHO 90", "JOELHO 45", "TE PVC", "TÊ PVC", "LUVA PVC", "ADAPTADOR PVC", "CURVA PVC", "JUNCAO PVC", "JUNÇÃO PVC", "BUCHA REDUCAO", "BUCHA REDUÇÃO", "CAP PVC", "PLUG PVC", "REGISTRO DE GAVETA", "REGISTRO DE PRESSAO", "REGISTRO ESFERA", "VALVULA DE RETENCAO", "VÁLVULA DE RETENÇÃO", "RALO SIFONADO", "CAIXA SIFONADA", "CAIXA DE GORDURA", "CAIXA D AGUA", "CAIXA D'AGUA", "BOIA", "FITA VEDA ROSCA", "ADESIVO PLASTICO PVC", "ADESIVO PLÁSTICO PVC", "SOLUCAO LIMPADORA", "SOLUÇÃO LIMPADORA", "PASTA LUBRIFICANTE", "TUBO DE COBRE PARA REFRIGERA", "ISOLANTE ESPONJOSO", "CONFORTO ACUSTICO", "CONFORTO ACÚSTICO", "VALVULA RETENCAO", "VÁLVULA RETENÇÃO"]):
        return "INSTALAÇÕES HIDRÁULICAS"

    # 3. Instalações Elétricas
    if any(k in d for k in ["CABO FLEXIVEL", "CABO FLEXÍVEL", "CABO DE COBRE", "CABO PP", "FIO COPEL", "DISJUNTOR", "INTERRUPTOR", "TOMADA", "BARRAMENTO", "LUMINARIA", "LUMINÁRIA", "REFLETOR", "LAMPADA", "LÂMPADA", "LED", "PLAFON", "QUADRO DE DISTRIBUI", "CAIXA DE EMBUTIR PVC", "ELETRODUTO", "CONDUITE", "CONDULETE", "PERFILADO", "LEITO", "DPS", "DR BIPOLAR", "FITA ISOLANTE", "PLUG MACHO", "PLUG FEMEA", "PLUG FÊMEA", "CONECTOR GENERICO", "CONECTOR GENÉRICO", "PAINEL LED", "FITA LED", "CURVA 90 PVC RIGIDA", "CURVA 90° PVC RÍGIDA", "ANILHAS DE IDENTIFICA"]):
        return "INSTALAÇÕES ELÉTRICAS"

    # 4. Argamassas e Rejuntes
    if any(k in d for k in ["ARGAMASSA", "CHAPISCO COLANTE", "CHAPISCO ROLADO", "REBOCO", "REJUNTE", "GRAUTH", "GRAUTE", "SUPERGRAUTE"]):
        return "ARGAMASSAS"

    # 5. Agregados e Aglomerantes
    if any(k in d for k in ["CIMENTO PORTLAND", "CIMENTO CP", "AREIA MEDIA", "AREIA MÉDIA", "AREIA GROSSA", "AREIA FINA", "BRITA 0", "BRITA 1", "BRITA 2", "BRITA GRADUADA", "CAL HIDRATADA", "GESSO EM PO", "GESSO EM PÓ", "SAIBRO", "PEDRA MARROADA"]):
        return "AGREGADOS E AGLOMERANTES"

    # 6. Blocos e Tijolos
    if any(k in d for k in ["TIJOLO", "BLOCO CERAMICO", "BLOCO CERÂMICO", "BLOCO DE CONCRETO", "BLOCO ESTRUTURAL", "ELEMENTO VAZADO", "COBOGO", "COBOGÓ", "CANALETA CERAMICA", "CANALETA DE CONCRETO"]):
        return "BLOCOS E TIJOLOS"

    # 7. Madeira
    if any(k in d for k in ["TABUA MISTA", "TÁBUA MISTA", "BARROTE", "CAIBRO", "SARRAFO", "VIGA DE MADEIRA", "COMPENSADO PLASTIFICADO", "COMPENSADO RESINADO", "COMPENSADO 12MM", "COMPENSADO 14MM", "COMPENSADO 18MM", "PONTALETE", "MADEIRA MISTA", "PINUS", "CHAPA RECICLAVEL", "CHAPA RECICLÁVEL"]):
        return "MADEIRA"

    # 8. Impermeabilizantes e Isolamento
    if any(k in d for k in ["MANTA ASFALTICA", "MANTA ASFÁLTICA", "MANTEX", "BIANCO", "VIAFIX", "EMULSAO ASFALTICA", "EMULSÃO ASFÁLTICA", "VEDACIT", "SIKA", "DENVER", "COMPOUND", "ADESIVO TIX", "HIDROREPELENTE", "SILICONE ACETICO", "SILICONE NEUTRO", "SILICONE PU", "SELANTE DE SILICONE", "SELANTE PU", "PU 40", "PU CONSTRUCAO", "ISOPOR", "EPS", "MACROFIBRA", "MICROFIBRA", "FIBERSTRAND", "DESMOLDANTE", "DESFORMA"]):
        return "IMPERMEABILIZANTES"

    # 9. Ferramentas
    if any(k in d for k in ["DISCO DE CORTE", "DISCO DIAMANTADO", "DISCO DE LIXA", "DISCO DE VIDEA", "TALHADEIRA", "PONTEIRO SDS", "DESEMPENADEIRA", "RODEL", "COLHER DE PEDREIRO", "MARRETA", "MARTELO", "ESQUADRO", "TRENA", "PRUMO", "NIVEL DE ALUMINIO", "NÍVEL", "ALICATE", "TORQUES", "TORQUÊS", "CHAVE DE FENDA", "CHAVE PHILIPS", "CHAVE COMBINADA", "APLICADOR DE SILICONE", "PISTOLA DE SILICONE", "RISCADEIRA", "SERRA CIRCULAR", "SERRA MARMORE", "SERRA MÁRMORE", "ESMERILHADEIRA", "FURADEIRA", "PARAFUSADEIRA", "FERRO DE SOLDA", "MANGUEIRA DE NIVEL", "ESPATULA", "ESPÁTULA", "BROCA SDS", "BROCA DE WIDEA", "BROCA DE ACO", "BROCA DE AÇO", "BROCA AÇO RÁPIDO", "BROCA ACO RAPIDO"]):
        return "FERRAMENTAS"

    # 10. Tintas e Pintura
    if any(k in d for k in ["TINTA ESMALTE", "TINTA ACRILICA", "TINTA ACRÍLICA", "TINTA LATEX", "TINTA LÁTEX", "TINTA SPRAY", "VERNIZ", "SELADOR", "FUNDO PREPARADOR", "SOLVENTE", "THINNER", "AGUARRAS", "AGUARRÁS", "MASSA CORRIDA", "MASSA ACRILICA", "MASSA ACRÍLICA", "ROLO DE LA", "ROLO DE LÃ", "ROLO DE ESPUMA", "PINCEL", "TRINCHA", "BROXA", "FITA CREPE", "LIXA MASSA", "LIXA FERRO"]):
        return "TINTAS"

    # 11. EPIs e EPCs
    if any(k in d for k in ["BOTINA", "BOTA DE COURO", "CAPACETE", "LUVA VAQUETA", "LUVA NITRILICA", "LUVA NITRÍLICA", "LUVA PIGMENTADA", "LUVA LATEX", "LUVA LÁTEX", "OCULOS DE PROTE", "ÓCULOS DE PROTE", "PROTETOR AURICULAR", "ABAFADOR", "CINTO DE SEGURAN", "TALABARTE", "TRAVA QUEDA", "CORDA SEG", "CORDA POLIAMIDA", "TELA TAPUME", "TELA FACHADA", "TELA GUARDA CORPO", "CAMISA HELANCA", "CALCA HELANCA", "CALÇA HELANCA", "PROTETOR DE CORDA", "AVENTAL", "MASCARA PFF", "MÁSCARA PFF", "RESPIRADOR", "PROTETOR SOLAR", "CONE DE SINALIZA", "CORRENTE PLASTICA", "CORRENTE PLÁSTICA"]):
        return "EPI'S E EPC'S"

    # 12. Material de Limpeza
    if any(k in d for k in ["DETERGENTE", "SABAO", "SABÃO", "AGUA SANITARIA", "ÁGUA SANITÁRIA", "DESINFETANTE", "SACO PLASTICO PARA LIXO", "SACO DE LIXO", "VASSOURA", "PANO DE CHAO", "PANO DE CHÃO", "PASTILHA DE CLORO", "CLORO", "LIMPADOR POS OBRA", "LIMPADOR PÓS OBRA", "DESENGRAXANTE", "INSETICIDA", "LUSTRA MOVEIS", "DESODORIZADOR", "SODA CAUSTICA", "SODA CÁUSTICA"]) or (re.search(r'\bRODO\b', d) and "ELETRODO" not in d):
        return "MATERIAL DE LIMPEZA"

    # 13. Material de Expediente
    if any(k in d for k in ["PAPEL OFICIO", "PAPEL OFÍCIO", "PAPEL A4", "SULFITE", "TONER", "REFIL EPSON", "TINTA PARA IMPRESSORA", "CADERNO", "CANETA", "LAPIS", "LÁPIS", "PASTA AZ", "PORTA CNPJ", "FILME PARA PLASTIFICA", "PLASTIFICACAO", "GRAMPEADOR", "PERFURADOR", "NOTEBOOK", "COMPUTADOR", "MOUSE", "TECLADO", "CALCULADORA", "PILHA"]):
        return "MATERIAL DE EXPEDIENTE"

    # 14. Instalações de Incêndio
    if any(k in d for k in ["MANGUEIRA DE INCENDIO", "MANGUEIRA DE INCÊNDIO", "ABRIGO MET", "ABRIGO DE INCENDIO", "EXTINTOR", "ESGUICHO", "RANHURADO", "GROOVED", "ADAPTADOR STORZ", "VALVULA DE GOVERNO", "VALVULA GLOBO ANGULAR"]):
        return "INSTALAÇÕES DE INCÊNDIO"

    # 15. Instalações de Gás
    if any(k in d for k in ["TUBO COMPACT ALUMINIO FLEX", "TUBO COMPACT ALUMÍNIO", "TUBO MULTICAMADA GAS", "TUBO COBRE GAS", "REGULADOR DE GAS", "REGULADOR DE GÁS", "FLEXIVEL PARA GAS", "FLEXÍVEL PARA GÁS"]):
        return "INSTALAÇÕES DE GÁS"

    # 16. Louças, Metais e Acessórios
    if any(k in d for k in ["BACIA SANITARIA", "BACIA SANITÁRIA", "BACIA LOUCA", "BACIA LOUÇA", "VASO SANITARIO", "VASO SANITÁRIO", "LAVATORIO", "LAVATÓRIO", "CUBA", "MICTORIO", "MICTÓRIO", "CAIXA DE DESCARGA", "SIFAO", "SIFÃO", "TORNEIRA", "VALVULA DE ESCOAMENTO", "VÁLVULA DE ESCOAMENTO", "ASSENTO SANITARIO", "DUCHA HIGIENICA", "DUCHA HIGIÊNICA"]):
        return "LOUÇAS E METAIS"

    # 17. Pavimentação e Drenagem
    if any(k in d for k in ["TUBO PEAD", "GEOCAMISA", "MANTA GEOTEXTIL", "MANTA GEOTÊXTIL", "TUBO DRENO", "GRELHA DRENAGEM", "PISO INTERTRAVADO", "PAVER", "MEIO FIO", "MEIO-FIO"]):
        return "PAVIMENTAÇÃO E DRENAGEM"

    # 18. Ferragens
    if any(k in d for k in ["DOBRADICA", "DOBRADIÇA", "FECHADURA", "TRINCO", "CADEADO", "FERROLHO", "ROLDANA", "PUXADOR", "PARAFUSO PARA DRYWALL", "PARAFUSO CHIPBOARD", "BUCHA NYLON", "BUCHA S6", "BUCHA S8", "BUCHA S10"]):
        return "FERRAGENS"

    # 19. Serviços Empreitados e Locação
    if any(k in d for k in ["SERVICO DE", "SERVIÇO DE", "SERVICOS DE", "SERVIÇOS DE", "LOCACAO DE", "LOCAÇÃO DE", "ALUGUEL DE", "MONTAGEM DE BALANCIM", "DESMONTAGEM DE BALANCIM", "CORTE E DOBRA", "FORRO DE GESSO"]):
        return "SERVIÇOS EMPREITADOS"

    # 20. Equipamentos
    if any(k in d for k in ["BETONEIRA", "GERADOR", "GUINCHO", "ELEVADOR", "COMPACTADOR", "ROMPEDOR", "BOMBA DE CONCRETO", "PALETEIRA", "CARRINHO DE MAO", "CARRINHO DE MÃO", "CARRO ARMAZEM", "CARRO ARMAZÉM", "AR CONDICIONADO", "VENTILADOR DE PAREDE", "BEBEDOURO", "MAQUINA DE ASSENTAR", "MÁQUINA DE ASSENTAR", "MAQUINA DE CORTE", "MÁQUINA DE CORTE", "CORTADOR ELETRICO", "CORTADOR ELÉTRICO", "ZAPP", "ARMARIO", "ARMÁRIO", "ROPEIRO", "ROUPEIRO"]) or c in ["13876", "10298"]:
        return "EQUIPAMENTOS"

    # 21. Materiais Auxiliares Específicos
    if any(k in d for k in ["ESPUMA COLCHAO", "ESPUMA COLCHÃO", "COLCHAO LAMINADO", "COLCHÃO LAMINADO", "ESPUMA"]) or c in ["10617"]:
        return "MATERIAIS AUXILIARES"

    # Fallback: limpar rótulo existente
    if current_fam and "DIVERSOS" not in str(current_fam).upper() and "SEM GRUPO" not in str(current_fam).upper():
        clean = re.sub(r'^\d+(\.\d+)*\s*[-–\.]*\s*', '', str(current_fam)).strip().upper()
        return CANONICAL_FAMILIES.get(clean, clean)

    return "SEM GRUPO"


def clean_family_label(fam: str, desc: str = "", cod: str = "") -> str:
    if desc or cod:
        return classify_item(desc, cod, fam)
    if not fam:
        return "SEM GRUPO"
    fam = str(fam).strip()
    fam = re.sub(r'^\d+(\.\d+)*\s*[-–\.]*\s*', '', fam).strip().upper()
    return CANONICAL_FAMILIES.get(fam, fam or "SEM GRUPO")


def parse_date(d_str: Any) -> Optional[datetime]:
    if not d_str:
        return None
    d_str = str(d_str).strip()[:10]
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(d_str, fmt)
        except Exception:
            pass
    return None


def format_brl(val: float) -> str:
    try:
        val_f = float(val or 0.0)
        formatted = f"{val_f:,.2f}"
        return "R$" + formatted.replace(',', 'X').replace('.', ',').replace('X', '.')
    except Exception:
        return "R$0,00"



def format_qty(val: Any) -> str:
    try:
        val_f = float(val or 0.0)
        if val_f.is_integer():
            return f"{int(val_f):,}".replace(',', '.')
        else:
            return f"{val_f:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except Exception:
        return str(val or "")


def get_previous_month_start(ref_dt: Optional[datetime] = None) -> datetime:
    if ref_dt is None:
        ref_dt = datetime.today()
    primeiro_dia_atual = ref_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if primeiro_dia_atual.month == 1:
        return primeiro_dia_atual.replace(year=primeiro_dia_atual.year - 1, month=12)
    else:
        return primeiro_dia_atual.replace(month=primeiro_dia_atual.month - 1)


def get_week_month_info(dt: datetime) -> tuple[int, int, int, str, str, datetime, datetime]:
    """
    Calcula a semana com reset a cada início de mês.
    Retorna: (ano_ref, mes_ref, sem_no_mes, label_semana, periodo_str, inicio_sem, fim_sem)
    """
    inicio_sem = dt - timedelta(days=dt.weekday())
    inicio_sem = inicio_sem.replace(hour=0, minute=0, second=0, microsecond=0)
    fim_sem = inicio_sem + timedelta(days=6)

    meio_sem = inicio_sem + timedelta(days=3)
    mes_ref = meio_sem.month
    ano_ref = meio_sem.year

    primeiro_dia_mes = datetime(ano_ref, mes_ref, 1)
    primeira_segunda = primeiro_dia_mes - timedelta(days=primeiro_dia_mes.weekday())
    if primeira_segunda + timedelta(days=3) < primeiro_dia_mes:
        primeira_segunda += timedelta(days=7)

    diff_dias = (inicio_sem - primeira_segunda).days
    sem_no_mes = (diff_dias // 7) + 1
    if sem_no_mes < 1:
        sem_no_mes = 1

    label = f"Sem {sem_no_mes:02d} - {NOMES_MESES[mes_ref]}/{ano_ref}"
    periodo = f"{inicio_sem.strftime('%d/%m')} a {fim_sem.strftime('%d/%m/%Y')}"
    return (ano_ref, mes_ref, sem_no_mes, label, periodo, inicio_sem, fim_sem)


def calculate_installments_for_item(record: Dict[str, Any]) -> List[Dict[str, Any]]:
    pc = str(record.get("numero_pedido", ""))
    dt_ped_str = record.get("data_pedido", "")
    dt_ent_str = record.get("data_entrega_prevista", "")
    fornec = str(record.get("fornecedor_nome", "")).split(" - ")[0][:32]
    cnpj = str(record.get("fornecedor_cnpj", "")).strip()
    fam = str(record.get("familia_insumo", "04 DIVERSOS"))
    desc = str(record.get("descricao_material", ""))[:45]
    val_total = float(record.get("preco_total_item", 0.0) or 0.0)
    cond_str = str(record.get("condicao_pagamento", "") or "").strip()
    cond_clean = cond_str.upper()

    dt_ped = parse_date(dt_ped_str) or datetime.today()
    dt_ent = parse_date(dt_ent_str) or dt_ped
    hoje = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    installments = []

    def make_inst_dict(num_p, val_p, dt_venc, regra):
        diff_dias = (dt_venc.replace(hour=0, minute=0, second=0, microsecond=0) - hoje).days
        if diff_dias < 0:
            st = f"Vencido ({abs(diff_dias)}d atrás)"
        elif diff_dias == 0:
            st = "Vence Hoje"
        elif diff_dias <= 7:
            st = f"Vence em {diff_dias}d (Esta Semana)"
        else:
            st = f"No Prazo (+{diff_dias}d)"

        ano_r, mes_r, sem_m, semana_str, periodo_str, inicio_s, fim_s = get_week_month_info(dt_venc)
        mes_ano_str = f"{NOMES_MESES[dt_venc.month]}/{dt_venc.year}"

        return {
            "numero_pedido": pc,
            "data_pedido": dt_ped.strftime("%d/%m/%Y"),
            "data_entrega": dt_ent.strftime("%d/%m/%Y"),
            "fornecedor": fornec,
            "cnpj": cnpj,
            "familia": fam,
            "material": desc,
            "condicao_original": cond_str or "Não informada",
            "parcela": num_p,
            "regra_prazo": regra,
            "valor_parcela": val_p,
            "data_vencimento": dt_venc.strftime("%d/%m/%Y"),
            "mes_ano": mes_ano_str,
            "semana_ano": semana_str,
            "status_prazo": st,
            "_venc_dt": dt_venc,
            "_ano_ref": ano_r,
            "_mes_ref": mes_r,
            "_sem_no_mes": sem_m,
            "_periodo_str": periodo_str,
            "_inicio_sem": inicio_s,
            "_fim_sem": fim_s
        }


    # 1. 50% Antecipado e 50% na Entrega
    if "50%" in cond_clean and "ENTREGA" in cond_clean:
        installments.append(make_inst_dict("1/2", val_total * 0.5, dt_ped, "50% Sinal na Emissão"))
        installments.append(make_inst_dict("2/2", val_total * 0.5, dt_ent, "50% Saldo na Entrega"))
        return installments

    # 2. 25% Antecipado + 30/60/90
    if "25%" in cond_clean:
        installments.append(make_inst_dict("1/4", val_total * 0.25, dt_ped, "25% Sinal na Emissão"))
        saldo_parc = (val_total * 0.75) / 3.0
        for i, d in enumerate([30, 60, 90]):
            installments.append(make_inst_dict(f"{i+2}/4", saldo_parc, dt_ent + timedelta(days=d), f"+{d} dias entrega"))
        return installments

    # 3. À Vista / Pix / Antecipado Total
    if any(k in cond_clean for k in ["AVISTA", "A VISTA", "PIX", "DINHEIRO", "ANTECIPADO"]):
        base_dt = dt_ped if "ANTECIPADO" in cond_clean else dt_ent
        reg = "Pagamento na Emissão" if "ANTECIPADO" in cond_clean else "À Vista na Entrega"
        installments.append(make_inst_dict("1/1 (À Vista)", val_total, base_dt, reg))
        return installments

    # 4. Prazos Numéricos (ex: 30 / 60 / 90 DIAS ou 28 DIAS)
    nums = [int(n) for n in re.findall(r'\b\d+\b', cond_clean)]
    if nums:
        n_parc = len(nums)
        v_parc = val_total / n_parc
        for i, d in enumerate(nums):
            installments.append(make_inst_dict(f"{i+1}/{n_parc}", v_parc, dt_ent + timedelta(days=d), f"+{d} dias"))
        return installments

    # Fallback: 30 dias após entrega
    installments.append(make_inst_dict("1/1", val_total, dt_ent + timedelta(days=30), "+30 dias entrega"))
    return installments


def load_familias_catalog() -> Dict[str, str]:
    candidate_paths = [
        r"G:\Meu Drive\Maison Plage - Pedidos de Compra\Configurações e Catálogos\familias_insumos.json",
        r"G:\My Drive\Maison Plage - Pedidos de Compra\Configurações e Catálogos\familias_insumos.json",
        os.path.abspath("data/familias_insumos.json")
    ]
    for p in candidate_paths:
        if os.path.exists(p):
            try:
                with open(p, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    return {k: clean_family_label(v.get("familia_original") or v.get("familia", "")) for k, v in data.items()}
            except Exception:
                pass
    return {}


def load_verified_suppliers() -> Dict[str, Dict[str, str]]:
    candidate_paths = [
        r"G:\Meu Drive\Maison Plage - Pedidos de Compra\Configurações e Catálogos\fornecedores_verificados.json",
        r"G:\My Drive\Maison Plage - Pedidos de Compra\Configurações e Catálogos\fornecedores_verificados.json",
        os.path.abspath("data/fornecedores_verificados.json")
    ]
    for p in candidate_paths:
        if os.path.exists(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
    return {}


def clean_supplier_name(raw_name: str, cnpj: str = "") -> str:
    cnpj_clean = str(cnpj or "").strip()
    catalog = load_verified_suppliers()
    if cnpj_clean in catalog:
        return catalog[cnpj_clean]["nome"]

    raw = str(raw_name or "").strip()
    if not raw or re.match(r'^\d{5}\-?\d{3}$', raw):
        return "FORNECEDOR DIVERSOS"

    # Remove CNPJ e telefones grudados no nome
    raw = re.sub(r'CNPJ[\d\.\/\-]+', '', raw, flags=re.IGNORECASE)
    raw = re.sub(r'Telefone.*', '', raw, flags=re.IGNORECASE)
    raw = re.sub(r'^\d{2}\.\d{3}\.\d{3}\/\d{4}\-\d{2}\s*', '', raw)
    raw = re.sub(r'^\d+\s*[-–]\s*', '', raw)

    # Deduplicação de repetições consecutivas
    words = raw.split()
    if len(words) >= 4:
        metade = len(words) // 2
        if words[:metade] == words[metade:2*metade]:
            words = words[:metade]
    raw = " ".join(words)
    raw = re.sub(r'\s+', ' ', raw).strip()

    return raw or "FORNECEDOR DIVERSOS"


def check_prazo_status(data_entrega_str: str) -> tuple[str, str]:
    if not data_entrega_str:
        return ("Sem Data Prevista", "pendente")

    dt_entrega = parse_date(data_entrega_str)
    if not dt_entrega:
        return ("Data Inválida", "pendente")

    hoje = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    diff = (dt_entrega - hoje).days

    if diff < 0:
        return (f"Atrasado ({abs(diff)} dias)", "atrasado")
    elif diff == 0:
        return ("Previsto para Hoje", "hoje")
    elif diff <= 7:
        return (f"Esta Semana ({diff} dias)", "proximo")
    elif diff <= 15:
        return (f"Próxima Semana ({diff} dias)", "no_prazo")
    else:
        return (f"No Prazo ({diff} dias)", "no_prazo")


class ExcelManager:
    def __init__(self, file_path: str = "data/pedidos_compra_consolidado.xlsx"):
        self.file_path = os.path.abspath(file_path)
        self.familias_map = load_familias_catalog()
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

    def load_existing_records(self) -> Dict[str, Dict[str, Any]]:
        records = {}
        if not os.path.exists(self.file_path):
            return records

        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb["Pedidos de Compra"] if "Pedidos de Compra" in wb.sheetnames else wb.active
            headers = [cell.value for cell in ws[1]]
            header_map = {name: idx for idx, name in enumerate(headers) if name}

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue

                def get_val(col_name):
                    idx = header_map.get(col_name)
                    return row[idx] if idx is not None and idx < len(row) else ""


                num_ped = str(get_val("Nº Pedido") or "").strip()
                cod_insumo = str(get_val("Cód. Insumo") or "").strip()
                desc = str(get_val("Descrição do Material / Insumo") or "").strip()

                if not num_ped or (not cod_insumo and not desc):
                    continue

                familia = classify_item(desc, cod_insumo, get_val("Família Sienge") or self.familias_map.get(cod_insumo, ""))

                key = f"{num_ped}_{row_idx}_{cod_insumo or desc}"
                records[key] = {
                    "numero_pedido": num_ped,
                    "data_pedido": get_val("Data do Pedido"),
                    "fornecedor_nome": get_val("Fornecedor"),
                    "fornecedor_cnpj": get_val("CNPJ Fornecedor"),
                    "codigo_insumo": cod_insumo,
                    "familia_insumo": familia,
                    "descricao_material": desc,
                    "quantidade": get_val("Quantidade"),
                    "unidade": get_val("Unid."),
                    "preco_unitario": get_val("Pr. Unitário (R$)"),
                    "preco_total_item": get_val("Total Item (R$)"),
                    "data_entrega_prevista": get_val("Entrega Prevista"),
                    "condicao_pagamento": get_val("Condição Pagamento"),
                    "situacao_prazo": get_val("Previsão / Prazo") or get_val("Situação do Prazo") or "",
                    "observacoes_sienge": get_val("Observações Sienge") or "",
                    "arquivo_origem": get_val("Arquivo Fonte") or "",
                    "ultima_atualizacao": get_val("Última Atualização") or "",
                    "valor_frete": get_val("Frete Pedido (R$)") or 0.0,
                    "valor_total_pedido": get_val("Total Pedido Sienge (R$)") or 0.0
                }
            wb.close()
        except Exception as e:
            print(f"Aviso ao ler planilha: {e}")

        return records

    def sync_extracted_orders(self, extracted_orders: List[Dict[str, Any]]) -> int:
        existing_records = self.load_existing_records()
        from zoneinfo import ZoneInfo
        now_str = datetime.now(ZoneInfo('America/Sao_Paulo')).strftime("%d/%m/%Y %H:%M:%S")

        for order in extracted_orders:
            num_ped = str(order.get("numero_pedido") or "").strip()
            
            # Remove todos os itens antigos deste pedido antes de reinserir (Evita duplicação/fantasmas em aditivos)
            if num_ped:
                keys_to_delete = [k for k, v in existing_records.items() if str(v.get("numero_pedido", "")).strip() == num_ped]
                for k in keys_to_delete:
                    del existing_records[k]

            data_ped = str(order.get("data_pedido") or "").strip()
            fornec_info = order.get("fornecedor", {})
            fornec_nome = fornec_info.get("nome", "")
            fornec_cnpj = fornec_info.get("cnpj", "")
            condicoes = order.get("condicoes", {})
            data_entrega = condicoes.get("data_entrega", "")
            cond_pagto = condicoes.get("condicao_pagamento", "")
            obs_sienge = order.get("observacoes", "")
            arquivo_origem = order.get("arquivo_origem", "")
            val_frete = float(condicoes.get("valor_frete", 0.0) or 0.0)
            val_total_sienge = float(order.get("total_pedido", 0.0) or 0.0)

            itens = order.get("itens", [])
            for item_idx, item in enumerate(itens, start=1):
                cod_insumo = str(item.get("codigo") or "").strip()
                desc_material = item.get("descricao") or item.get("descricao_completa") or ""
                key = f"{num_ped}_{item_idx}_{cod_insumo or desc_material}"

                familia = classify_item(desc_material, cod_insumo, self.familias_map.get(cod_insumo, ""))
                situacao_prazo, _ = check_prazo_status(data_entrega)

                record = {
                    "numero_pedido": num_ped,
                    "data_pedido": data_ped,
                    "fornecedor_nome": fornec_nome,
                    "fornecedor_cnpj": fornec_cnpj,
                    "codigo_insumo": cod_insumo,
                    "familia_insumo": familia,
                    "descricao_material": desc_material,
                    "quantidade": item.get("quantidade"),
                    "unidade": item.get("unidade", ""),
                    "preco_unitario": item.get("preco_unitario"),
                    "preco_total_item": item.get("preco_final"),
                    "data_entrega_prevista": data_entrega,
                    "condicao_pagamento": cond_pagto,
                    "situacao_prazo": situacao_prazo,
                    "observacoes_sienge": obs_sienge,
                    "arquivo_origem": arquivo_origem,
                    "ultima_atualizacao": now_str,
                    "valor_frete": val_frete,
                    "valor_total_pedido": val_total_sienge
                }
                existing_records[key] = record

        self._write_workbook(list(existing_records.values()))
        return len(existing_records)

    def remove_orders(self, numeros_pedido: List[str]) -> int:
        """Remove pedidos completamente da base (ex: cancelamentos)."""
        existing_records = self.load_existing_records()
        removidos = 0
        for num_ped in numeros_pedido:
            num_ped = str(num_ped).strip()
            if num_ped:
                keys_to_delete = [k for k in existing_records.keys() if k.startswith(f"{num_ped}_")]
                for k in keys_to_delete:
                    del existing_records[k]
                    removidos += 1
        
        if removidos > 0:
            self._write_workbook(list(existing_records.values()))
        return removidos

    def _write_workbook(self, records: List[Dict[str, Any]]):
        wb = openpyxl.Workbook()

        # Estilos Padrão
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill_blue = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_fill_green = PatternFill(start_color="1E7145", end_color="1E7145", fill_type="solid")
        header_fill_orange = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        header_fill_purple = PatternFill(start_color="5A2D81", end_color="5A2D81", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        regular_font = Font(name="Calibri", size=10)
        bold_font = Font(name="Calibri", size=10, bold=True)
        zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # -------------------------------------------------------------
        # ABA 1: PEDIDOS DE COMPRA
        # -------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Pedidos de Compra"
        ws1.row_dimensions[1].height = 28

        for col_idx, col_def in enumerate(COLUMNS_DEFINITION, start=1):
            cell = ws1.cell(row=1, column=col_idx, value=col_def["title"])
            cell.font = header_font
            cell.fill = header_fill_blue
            cell.alignment = header_align

        for row_idx, record in enumerate(records, start=2):
            ws1.row_dimensions[row_idx].height = 20
            use_zebra = (row_idx % 2 == 0)

            for col_idx, col_def in enumerate(COLUMNS_DEFINITION, start=1):
                val = record.get(col_def["key"])
                cell = ws1.cell(row=row_idx, column=col_idx)

                if col_def["key"] in ["quantidade", "preco_unitario", "preco_total_item", "valor_frete", "valor_total_pedido"]:
                    try:
                        cell.value = float(val) if val not in [None, ""] else 0.0
                    except Exception:
                        cell.value = val
                else:
                    cell.value = str(val) if val is not None else ""


                cell.font = regular_font
                cell.alignment = Alignment(horizontal=col_def["align"], vertical="center")
                cell.border = thin_border
                cell.number_format = col_def["format"]

                if use_zebra:
                    cell.fill = zebra_fill

        for col_idx, col_def in enumerate(COLUMNS_DEFINITION, start=1):
            col_letter = get_column_letter(col_idx)
            ws1.column_dimensions[col_letter].width = col_def["width"]

        ws1.freeze_panes = "A2"

        # -------------------------------------------------------------
        # CALCULO DAS PARCELAS FILTRADAS (DO MÊS ANTERIOR EM DIANTE)
        # -------------------------------------------------------------
        mes_corte = get_previous_month_start()

        all_installments = []
        for r in records:
            insts = calculate_installments_for_item(r)
            all_installments.extend(insts)

        installments_filtradas = [i for i in all_installments if i["_venc_dt"] >= mes_corte]
        installments_filtradas.sort(key=lambda x: x.get("_venc_dt") or datetime.min)
        total_periodo_desembolso = sum(i["valor_parcela"] for i in installments_filtradas) or 1.0

        # -------------------------------------------------------------
        # ABA 2: FLUXO DE DESEMBOLSO PREVISTO (ENXUTO E COM SEMANA RESETADA)
        # -------------------------------------------------------------
        ws2 = wb.create_sheet(title="Fluxo de Desembolso Previsto")
        ws2.row_dimensions[1].height = 28

        for col_idx, col_def in enumerate(COLUMNS_DESEMBOLSO_ENXUTO, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=col_def["title"])
            cell.font = header_font
            cell.fill = header_fill_green
            cell.alignment = header_align

        for row_idx, inst in enumerate(installments_filtradas, start=2):
            ws2.row_dimensions[row_idx].height = 20
            use_zebra = (row_idx % 2 == 0)

            for col_idx, col_def in enumerate(COLUMNS_DESEMBOLSO_ENXUTO, start=1):
                val = inst.get(col_def["key"])
                cell = ws2.cell(row=row_idx, column=col_idx)

                if col_def["key"] == "valor_parcela":
                    try:
                        cell.value = float(val) if val not in [None, ""] else 0.0
                    except Exception:
                        cell.value = val
                else:
                    cell.value = str(val) if val is not None else ""

                cell.font = regular_font
                cell.alignment = Alignment(horizontal=col_def["align"], vertical="center")
                cell.border = thin_border
                cell.number_format = col_def["format"]

                if use_zebra:
                    cell.fill = zebra_fill

        for col_idx, col_def in enumerate(COLUMNS_DESEMBOLSO_ENXUTO, start=1):
            col_letter = get_column_letter(col_idx)
            ws2.column_dimensions[col_letter].width = col_def["width"]

        ws2.freeze_panes = "A2"

        # -------------------------------------------------------------
        # ABA 3: RESUMO DE DESEMBOLSO POR SEMANA (RESETADA A CADA MÊS)
        # -------------------------------------------------------------
        ws3 = wb.create_sheet(title="Desembolso por Semana")
        ws3.row_dimensions[1].height = 28

        semana_cols = [
            {"title": "Semana do Mês", "width": 26, "align": "left"},
            {"title": "Período da Semana", "width": 28, "align": "center"},
            {"title": "Total de Parcelas", "width": 18, "align": "center"},
            {"title": "Qtd. Pedidos (PCs)", "width": 18, "align": "center"},
            {"title": "Total a Pagar na Semana (R$)", "width": 30, "align": "right", "format": "R$ #,##0.00"},
            {"title": "% do Total Geral", "width": 18, "align": "right", "format": "0.00%"}
        ]

        for col_idx, col_def in enumerate(semana_cols, start=1):
            cell = ws3.cell(row=1, column=col_idx, value=col_def["title"])
            cell.font = header_font
            cell.fill = header_fill_orange
            cell.alignment = header_align

        # Agrupamento semanal resetado a cada mês
        week_map = defaultdict(lambda: {"count": 0, "pedidos": set(), "total": 0.0, "label": "", "periodo": ""})
        for inst in installments_filtradas:
            k = (inst["_ano_ref"], inst["_mes_ref"], inst["_sem_no_mes"], inst["_inicio_sem"])
            week_map[k]["count"] += 1
            week_map[k]["pedidos"].add(inst["numero_pedido"])
            week_map[k]["total"] += inst["valor_parcela"]
            week_map[k]["label"] = inst["semana_ano"]
            week_map[k]["periodo"] = inst["_periodo_str"]

        sorted_weeks = sorted(week_map.keys(), key=lambda x: (x[0], x[1], x[2]))

        for row_idx, k in enumerate(sorted_weeks, start=2):
            ws3.row_dimensions[row_idx].height = 22
            wg = week_map[k]
            cnt = wg["count"]
            p_cnt = len(wg["pedidos"])
            tot = wg["total"]
            pct = tot / total_periodo_desembolso

            r_data = [wg["label"], wg["periodo"], cnt, p_cnt, tot, pct]
            for col_idx, val in enumerate(r_data, start=1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal=semana_cols[col_idx-1]["align"], vertical="center")
                if "format" in semana_cols[col_idx-1]:
                    cell.number_format = semana_cols[col_idx-1]["format"]

        # Linha Total Semanal
        tot_sem_row = len(sorted_weeks) + 2
        ws3.cell(row=tot_sem_row, column=1, value="TOTAL GERAL").font = bold_font
        ws3.cell(row=tot_sem_row, column=2, value=f"{len(sorted_weeks)} semanas").font = bold_font
        ws3.cell(row=tot_sem_row, column=3, value=len(installments_filtradas)).font = bold_font
        ws3.cell(row=tot_sem_row, column=4, value=len(set(i["numero_pedido"] for i in installments_filtradas))).font = bold_font
        cell_tot_sem = ws3.cell(row=tot_sem_row, column=5, value=total_periodo_desembolso)
        cell_tot_sem.font = bold_font
        cell_tot_sem.number_format = "R$ #,##0.00"
        cell_pct_sem = ws3.cell(row=tot_sem_row, column=6, value=1.0)
        cell_pct_sem.font = bold_font
        cell_pct_sem.number_format = "0.00%"

        for col_idx, col_def in enumerate(semana_cols, start=1):
            col_letter = get_column_letter(col_idx)
            ws3.column_dimensions[col_letter].width = col_def["width"]

        ws3.freeze_panes = "A2"

        # -------------------------------------------------------------
        # ABA 4: RESUMO DE DESEMBOLSO POR MÊS
        # -------------------------------------------------------------
        ws4 = wb.create_sheet(title="Desembolso por Mês")
        ws4.row_dimensions[1].height = 28

        mes_cols = [
            {"title": "Mês / Ano", "width": 24, "align": "left"},
            {"title": "Total de Parcelas", "width": 18, "align": "center"},
            {"title": "Qtd. Pedidos (PCs)", "width": 18, "align": "center"},
            {"title": "Total Previsto a Pagar (R$)", "width": 30, "align": "right", "format": "R$ #,##0.00"},
            {"title": "% do Total Geral", "width": 18, "align": "right", "format": "0.00%"}
        ]

        for col_idx, col_def in enumerate(mes_cols, start=1):
            cell = ws4.cell(row=1, column=col_idx, value=col_def["title"])
            cell.font = header_font
            cell.fill = header_fill_purple
            cell.alignment = header_align

        mes_map = {}
        for inst in installments_filtradas:
            dt_v = inst["_venc_dt"]
            k = (dt_v.year, dt_v.month, inst["mes_ano"])
            if k not in mes_map:
                mes_map[k] = {"count": 0, "pedidos": set(), "total": 0.0}
            mes_map[k]["count"] += 1
            mes_map[k]["pedidos"].add(inst["numero_pedido"])
            mes_map[k]["total"] += inst["valor_parcela"]

        sorted_meses = sorted(mes_map.keys(), key=lambda x: (x[0], x[1]))

        for row_idx, k in enumerate(sorted_meses, start=2):
            ws4.row_dimensions[row_idx].height = 22
            cnt = mes_map[k]["count"]
            p_cnt = len(mes_map[k]["pedidos"])
            tot = mes_map[k]["total"]
            pct = tot / total_periodo_desembolso

            r_data = [k[2], cnt, p_cnt, tot, pct]
            for col_idx, val in enumerate(r_data, start=1):
                cell = ws4.cell(row=row_idx, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal=mes_cols[col_idx-1]["align"], vertical="center")
                if "format" in mes_cols[col_idx-1]:
                    cell.number_format = mes_cols[col_idx-1]["format"]

        # Linha Total Mensal
        tot_mes_row = len(sorted_meses) + 2
        ws4.cell(row=tot_mes_row, column=1, value="TOTAL GERAL").font = bold_font
        ws4.cell(row=tot_mes_row, column=2, value=len(installments_filtradas)).font = bold_font
        ws4.cell(row=tot_mes_row, column=3, value=len(set(i["numero_pedido"] for i in installments_filtradas))).font = bold_font
        cell_tot_mes = ws4.cell(row=tot_mes_row, column=4, value=total_periodo_desembolso)
        cell_tot_mes.font = bold_font
        cell_tot_mes.number_format = "R$ #,##0.00"
        cell_pct_mes = ws4.cell(row=tot_mes_row, column=5, value=1.0)
        cell_pct_mes.font = bold_font
        cell_pct_mes.number_format = "0.00%"

        for col_idx, col_def in enumerate(mes_cols, start=1):
            col_letter = get_column_letter(col_idx)
            ws4.column_dimensions[col_letter].width = col_def["width"]

        ws4.freeze_panes = "A2"

        # Salva o arquivo final
        wb.save(self.file_path)
        wb.close()

        # Sincronização Google Drive (Desktop e Nuvem API)
        try:
            from src.gdrive_service import GoogleDriveSyncService
            gdrive = GoogleDriveSyncService()
            gdrive.sync_local_and_drive(self.file_path)
        except Exception as e:
            logger.warning(f"Aviso ao sincronizar com Google Drive: {e}")


# Alias para compatibilidade
ExcelDatabaseManager = ExcelManager
